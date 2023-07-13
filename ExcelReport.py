'''
Available methods are the followings:
[1] AutoReport
[2] ListValidation
[3] CellStyle

Authors: Danusorn Sitdhirasdr <danusorn.s@kasikornbank.com>
versionadded:: 11-07-2023

'''
import pandas as pd, numpy as np, os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import (PatternFill, Border, Side, 
                             Alignment, Font, Protection)
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting import Rule
from openpyxl.utils import get_column_letter
from itertools import product
from datetime import datetime

__all__ = ["AutoReport","ListValidation","CellStyle"]

def ListValidation(sh, source, coord=(1,1), 
                   offset=(0,0), kwargs=None):
    
    '''
    Add list validation to range.

    Parameters
    ----------
    sh : openpyxl Worksheet
        Excel worksheet.
        
    source : str, default=""
        
    coord : (int,int), default=(1,1)
        Starting cell coordinate e.g. (1,1) is "A1".
        
    offset : (int,int), default=(0,0)
        The number of rows, and columns to be added.
        
    kwargs: dict, default=None
        Other attributes i.e. 'error', 'errortitle', 'prompt', 
        and 'prompttitle'.
        
    References
    ----------
    [1] https://openpyxl.readthedocs.io/en/stable/styles.html
        
    '''
    # Create a data-validation object with list validation
    source = source[:255].lstrip().rstrip()
    dv = DataValidation(type="list", allow_blank=True, 
                        formula1=f'"{source}"')

    # Optionally set a custom message
    attrs = ["error", "errortitle", "prompt", "prompttitle"]
    if isinstance(kwargs, dict):
        for key in set(attrs).intersection(kwargs.keys()):
            setattr(dv, key, kwargs[key])
    
    # Add the data-validation object to designated cell
    cell0 = sh.cell(*coord).coordinate
    cell1 = sh.cell(coord[0] + max(offset[0], 0), 
                    coord[1] + max(offset[1], 0)).coordinate
    dv.add(f"{cell0}:{cell1}")
    sh.add_data_validation(dv)


# In[5]:


def UpdateDict(dict1:dict, dict2:dict):
    '''Update dictionary'''
    if isinstance(dict1, dict) & isinstance(dict2, dict):
        for key in set(dict1.keys()).intersection(dict2.keys()):
            dict1[key] = dict2[key]
    return dict1


# In[6]:


def CreateCoords(coord=(1,1), offset=(0,0)):
    '''Create coordinates'''
    stop = tuple(np.r_[coord] + offset + (1,1))
    return list(product(range(coord[0], stop[0]),
                        range(coord[1], stop[1])))


# In[7]:


def GetCellRange(sh, coord=(1,1), offset=(0,0)):
    '''Return address of cell range'''
    start = sh.cell(*coord).coordinate
    stop = tuple(np.r_[coord] + offset)
    stop = sh.cell(*stop).coordinate
    return start, stop


# In[8]:


class CellStyle:
    
    '''
    Change cell format.
    
    Parameters
    ----------
    kwargs: dict, default=None
        Other parameters of the following openpyxl.styles functions:
        - PatternFill e.g. "fill_type", "start_color"
        - Alignment e.g. "horizontal", "vertical"
        - Font e.g. "name", "underline"
        - Border e.g. "left", "right", whose value must be 
          "openpyxl.styles.Side(border_style, color)"
    
    References
    ----------
    [1] https://openpyxl.readthedocs.io/en/stable/styles.html
    
    '''
    def __init__(self, kwargs=None):
        
        # Default properties
        # underline = {'single', 'double', 
        #              'doubleAccounting', 
        #              'singleAccounting'}
        self.fontstyle = dict(name='Tahoma', 
                              size=10, 
                              bold=False, 
                              italic=False, 
                              vertAlign=None, 
                              underline='none', 
                              strike=False, 
                              color='404040')

        self.fillstyle = dict(fill_type="none", 
                              start_color='000000', 
                              end_color='000000', 
                              patternType='none')
        
        self.alignment = dict(horizontal='general',
                              vertical='bottom',
                              text_rotation=0,                 
                              wrap_text=False,
                              shrink_to_fit=False,                    
                              indent=0)
        
        # vertical and horizontal are excluded
        # same parameters as "alignment"
        sidestyle = Side(border_style=None, color='000000')
        self.borderstyle = dict(left=sidestyle,
                                right=sidestyle,
                                top=sidestyle,
                                bottom=sidestyle,
                                diagonal=sidestyle,
                                diagonal_direction=0,
                                outline=sidestyle)
        
        kwargs = dict() if kwargs is None else kwargs
        self.fontstyle = Font(**UpdateDict(self.fontstyle, kwargs))
        self.fillstyle = PatternFill(**UpdateDict(self.fillstyle, kwargs))
        self.alignment = Alignment(**UpdateDict(self.alignment, kwargs))
        self.borderstyle = Border(**UpdateDict(self.borderstyle, kwargs))
    
    def transform(self, sh, coord=(1,1), offset=(0,0)):
        
        '''
        Transform designated cell(s) according to predefined format.
        
        Parameters
        ----------
        sh : openpyxl Worksheet
            Excel worksheet.

        coord : (int,int), default=(1,1)
            Starting cell coordinate e.g. (1,1) is "A1".

        offset : (int,int), default=(0,0)
            The number of rows, and columns to be added.
            
        '''
        for cell in CreateCoords(coord, offset):
            addr = sh.cell(*cell).coordinate
            sh[addr].font = self.fontstyle
            sh[addr].fill = self.fillstyle
            sh[addr].alignment = self.alignment
            sh[addr].border = self.borderstyle


# In[9]:


class IsBlankformat(CellStyle):
    
    '''
    Change cell format when blank.
    
    Parameters
    ----------
    kwargs: dict, default=None
        Other parameters of the following openpyxl.styles functions:
        - PatternFill e.g. "fill_type", "start_color"
        - Alignment e.g. "horizontal", "vertical"
        - Font e.g. "name", "underline"
        - Border e.g. "left", "right", whose value must be 
          "openpyxl.styles.Side(border_style, color)"
    
    References
    ----------
    [1] https://openpyxl.readthedocs.io/en/stable/formatting.html
    
    '''
    def __init__(self, kwargs=None):
        
        # alignment is excluded (not compatible)
        self.pattern = CellStyle(kwargs)
        self.params = {"font"   : self.pattern.fontstyle, 
                       "fill"   : self.pattern.fillstyle, 
                       "border" : self.pattern.borderstyle}

    def transform(self, sh, coord=(1,1), offset=(0,0)):
        
        '''
        Add conditional format (blank) to designated cell(s).
        
        Parameters
        ----------
        sh : openpyxl Worksheet
            Excel worksheet.

        coord : (int,int), default=(1,1)
            Starting cell coordinate e.g. (1,1) is "A1".

        offset : (int,int), default=(0,0)
            The number of rows, and columns to be added.
            
        '''
        addr = GetCellRange(sh, coord, offset)
        rule = FormulaRule(formula=[f'ISBLANK({addr[0]})'], 
                           stopIfTrue=True, **self.params)
        sh.conditional_formatting.add(":".join(addr), rule)


# In[10]:


def GetColumn_base(wb, sheetname="columns"):
    '''Return orderly arranged columns'''
    sh = wb[sheetname]
    columns = [(sh.cell(n,1).value, 
                sh.cell(n,2).value, # Column name
                sh.cell(n,3).value, # Width
                sh.cell(n,4).value) # Color
               for n in range(2, sh.max_row + 1)]
    columns.sort(reverse=False, key=lambda n:n[0])
    return tuple(np.r_[columns][:,n] for n in range(1,4))


# In[11]:


def GetInput_base(wb, sheetname="inputs"):
    '''
    Return dict of inputs
    {
     "A": {"Col_A1": "a1,b1,c1,d1",
           "Col_A2": "a2,b2,c2,d2"},
     "B": {"Col_B1": "a1,b1,c1,d1",
           "Col_B2": "a2,b2,c2,d2"}
    }      
    '''
    sh = wb[sheetname]
    inputs = dict()
    for n in range(2, sh.max_row + 1):
        key = str(sh.cell(n,1).value)
        col = sh.cell(n,2).value
        val = sh.cell(n,3).value
        if key not in inputs.keys():
            inputs[key] = {}
        inputs[key].update({col: val})
    return inputs


# In[12]:


def GetParameters(path):
    '''
    Get all components i.e inputs, columns, width, and colors
    '''
    wb = load_workbook(path)
    sheets = ["inputs", "columns"]
    retval = (None, None)
    if sum(np.isin(wb.sheetnames, sheets)) == len(sheets):
        retval = (GetInput_base(wb),) + GetColumn_base(wb)
    wb.close()
    return retval


# In[13]:


def CopyExcel(source, sheetname):
    
    '''
    Copy worksheet to new worksheet.
    
    Parameters
    ----------
    source : str
        Source file path e.g. "C:\\Folder\\exmaple.xlsx"
        
    sheetname : str
        String that is used for sheet name.
        
    References
    ----------
    [1] https://openpyxl.readthedocs.io/en/stable/tutorial.html
    
    Returns
    -------
    wb2 : openpyxl.Workbook
    
    '''
    # opening the source excel file
    wb1 = load_workbook(source)
    sh1 = wb1.worksheets[0]

    # opening the destination excel file 
    wb2 = Workbook() 
    sh2 = wb2.active
    sh2.title = sheetname

    # copying the cell values from source 
    # excel file to destination excel file
    for i in range (1, sh1.max_row + 1):
        for j in range (1, sh1.max_column + 1):
            # writing the read value to destination excel file
            sh2.cell(i,j).value = sh1.cell(i, j).value

    wb1.close()
    return wb2


# In[14]:


def FindColumn(sh, colname):
    '''Return column index (case insensitive)'''
    columns = np.array([n.value.lower() for n in sh["1"]])
    return get_column_letter(np.argmax(columns==colname) + 1)


# In[18]:


def AutoReport(source1, source2, destination="", password="admin"):
    
    '''
    Create new report automatically.
    
    Parameters
    ----------
    source1 : str
        A valid string path for template (*.xlsx), which contains 
        worksheets i.e. "inputs", and "columns". A local file could 
        be: C://localhost/path/to/template.xlsx.
        
    source2 : str
        A valid string path for data (*.xlsx).
        
    destination : str
        A valid string path of repository (folder), where report will 
        be stored. A local folder could be: C://localhost/.
        
    password : str
        Password to protect worksheet.
    
    Returns
    -------
    filename : str
        A valid string path of report.
        
    '''
    # Get parameters 
    inputs, columns, width, colors = GetParameters(source1)
    
    # Create sheet in new workbook
    new_wb = CopyExcel(source2, f"DATA_{datetime.now():%Y%m%d}")
    sh = new_wb.worksheets[0]
    sh.row_dimensions[1].height = 70
    
    # Initialize parameters
    CellFormats = dict()
    sides = ["right", "left", "top", "bottom"]

    # General format
    side1 = Side(border_style="thin", color='2f3542')
    side1 = dict(product(sides,[side1]))
    CellFormats["general"] = CellStyle(side1)

    # Format for cells with data validation (ISBLANK)
    side2 = Side(border_style="thin", color='2f3542')
    kwds = dict(fill_type="lightUp", start_color='FC427B')
    kwds.update(dict(product(sides,[side2])))
    CellFormats["isblank"] = IsBlankformat(kwds)

    # Format for normal cells
    side3 = Side(border_style="thin", color='2f3542')
    kwds = dict(fill_type="lightUp", start_color='2ed573')
    kwds.update(dict(product(sides,[side3])))
    CellFormats["notblank"] = IsBlankformat(kwds)

    # Fill pattern for normal columns
    side4 = Side(border_style="thin", color='2f3542')
    kwds = dict(fill_type="solid", start_color='747d8c', 
                size=9, wrap_text=True, color="ffffff", 
                vertical="center", horizontal="center")
    kwds.update(dict(product(sides,[side4])))
    CellFormats["columns"] = CellStyle(kwds)

    # Fill pattern of added columns
    for start_color in np.unique(colors):
        kwds["start_color"] = start_color
        kwds["color"] = "2f3542"
        CellFormats[start_color] = CellStyle(kwds)
        
    # Add new columns
    usedcols, i = sh.max_column + 1, 0
    for c,col in enumerate(columns, usedcols):
        sh.cell(1,c).value = col.lstrip().rstrip()
        sh.column_dimensions[get_column_letter(c)].width = width[i]
        CellFormats[colors[i]].transform(sh, (1,c))
        i += 1
        
    # Apply cell formats i.e. general and column formats
    CellFormats["general"].transform(sh, (2,1), (sh.max_row-2, sh.max_column-1))
    CellFormats["columns"].transform(sh, (1,1), (0, usedcols-2))
    
    # Add data validation and cell format
    for r,cell in enumerate(sh[FindColumn(sh,"rule")][1:],2):
        for c,col in enumerate(columns, usedcols):
            source = inputs[str(cell.value)].get(col, None)
            if isinstance(source, str):
                ListValidation(sh, source, (r,c))
                CellFormats["isblank"].transform(sh, (r,c))
                if len(source.split(","))==1:
                    sh.cell(r,c).value = source
            else: CellFormats["notblank"].transform(sh, (r,c))
            sh.cell(r,c).protection = Protection(locked=False)
    
    # Add "auto_filter"
    filters = sh.auto_filter
    filters.ref = f"A1:{get_column_letter(sh.max_column)}{sh.max_row}"
    
    # Protect sheet
    # Note: False --> True (reverse)
    sh.protection = SheetProtection(sh, 
                                    formatRows=False, 
                                    formatColumns=False, 
                                    selectLockedCells=False, 
                                    selectUnlockedCells=False, 
                                    autoFilter=False,
                                    sort=False,
                                    password=password)
    
    # Save and close workbook
    sh.sheet_view.showGridLines = False
    filename = f"{destination}REPORT_{datetime.now():%Y%m%d%H%M%S}.xlsx"
    new_wb.save(filename)
    new_wb.close()
    
    return filename