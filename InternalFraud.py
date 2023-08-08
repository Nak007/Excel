'''
Available methods are the followings:
[1] AuditReport (class)
[2] SendReport (class)

Authors: Danusorn Sitdhirasdr <danusorn.s@kasikornbank.com>
versionadded:: 10-08-2023

'''
import pandas as pd, numpy as np, os, re
from datetime import datetime, timedelta
from functools import partial
from collections import namedtuple
from itertools import product
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (PatternFill, Border, Side, 
                             Alignment, Font, Protection)
from openpyxl.utils import get_column_letter
import hashlib, hmac, base64
import logging
import ExcelLib as EXCEL
import OutlookLib as OUTLOOK
from tqdm.notebook import tqdm_notebook
import ipywidgets as widgets
from IPython.display import display
import shutil, logging
import urllib

__all__ = ["AuditReport", 
           "SendReport"]

# Initialize log
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
subfolder = "logging"

# Create subfolder under current working directory
folder = os.path.join(os.getcwd(), subfolder)
if os.path.exists(folder)==False: os.makedirs(folder)
path = os.path.join(folder, f"ASOF_{datetime.now():%Y%m%d}.txt")

# Create format of file-handler.
fmt='%(asctime)s : %(levelname)s : %(message)s'
datefmt='%m/%d/%Y %I:%M:%S %p'
Formatter = logging.Formatter(fmt=fmt, datefmt=datefmt)

file_handler = logging.FileHandler(path)
file_handler.setFormatter(Formatter)

logger.addHandler(file_handler)
logger.propagate = False

def Worksheet(sh, r, c):
    return sh.cell(r,c).value

def hmac_hash(raw, key="admin", digestmod="sha1"):
    
    '''Hashing'''
    raw = raw.encode("utf-8")
    key = key.encode('utf-8')
    hashed = hmac.new(key, raw, digestmod=digestmod)
    hashed = base64.encodebytes(hashed.digest())
    hashed = hashed.decode('utf-8').rstrip('=\n')
    return hashed

class sum_namedtuple:
    
    '''Merge named tuples'''
    def __init__(self, typename="MergedTuples"):
        self.typename = typename
    
    def sum_classes(self, *args):
        a = sum(map(lambda t:t._fields, args), ())
        return namedtuple(self.typename, a)
    
    def merge(self, *args):
        return self.sum_classes(*args)(*sum(args,()))

def GetParameters(sh, sheetnames:list):
    
    '''
    Get all parameters as follows:
    - Data validation (dv)
    - Hex-color for columns (hex)
    - Columns for new sheets, and hashed columns (cols)
    '''
    # Initialize parameters
    datavalid, hexcolors, hashedcol = dict(), dict(), dict()
    usecols = dict([(n,[]) for n in sheetnames + ["hash", "mandate"]])
    columns = dict([(c.value,n) for n,c in enumerate(sh[1],1) 
                    if c.value is not None])

    ws = partial(Worksheet, sh)
    for r in np.arange(2, sh.max_row + 1):

        # Validation parameters
        if (ws(r,2) != "text") & (ws(r,2) is not None):
            kwargs = {"type"     : ws(r, columns["type"]), 
                      "operator" : ws(r, columns["operator"]),
                      "formula1" : ws(r, columns["formula1"]), 
                      "formula2" : ws(r, columns["formula2"])}
            datavalid[ws(r,1)] = EXCEL.Validation(**kwargs)

        # Hex-color code
        hexcolors[ws(r,1)] = {c : ws(r, columns[c]).replace("#","") 
                              for c in ["start_color","color"]} 
 
        # Keep columns
        for key in usecols.keys():
            if ws(r, columns[key])>0: 
                usecols[key] += [(ws(r, columns[key]),ws(r,1))] 
        
    # Sort column labels in list
    for key in usecols.keys(): 
        usecols[key].sort(key=lambda x:x[0],reverse=False)
    
    Params = namedtuple("Params", ("dv","hex","cols"))
    return Params(datavalid, hexcolors, usecols)

def ApplyFormats(sh, CellFormats, 
                 ColumnFormats=dict(), 
                 DataValidations=dict()):
    
    '''Apply format and data validation'''
    # General format for every cell
    columns = [c.value for c in sh[1]]
    CellFormats["general"].apply(sh, (1,1), 
                                 (max_rows := np.fmax(sh.max_row-1,1), 
                                  max_cols := len(columns)-1))
    CellFormats["isblank"].apply(sh, (2,1), (max_rows-1, max_cols))
    CellFormats["columns"].apply(sh, (1,1), (0, max_cols))

    # Apply data validation and column format
    for c,col in enumerate(columns, 1):
        
        # Apply column formats
        if col in ColumnFormats.keys(): 
            ColumnFormats[col].apply(sh, (1,c))
            
        # Apply data validation and default list value
        if col in DataValidations.keys(): 
            DataValidations[col].apply(sh, (2,c), (max_rows-1,0))
            if (DataValidations[col].dv.type=="list") & \
            (DataValidations[col].dv.formula1.find("ไม่ระบุ")>-1):
                for r in range(2, max_rows + 2):
                    sh.cell(r,c).value = "ไม่ระบุ"

def GeneralFormats(params):
    
    '''
    General formats for all cells as well as fill pattern
    for newly added columns.
    params : output from GetParameters
    '''
    # Initialize parameters
    CellFormats = dict()
    sides = ["right", "left", "top", "bottom"]

    # General format
    side1 = Side(border_style="thin", color='2f3542')
    side1 = dict(product(sides, [side1]))
    CellFormats["general"] = EXCEL.CellStyle(side1)

    # Format for cells with data validation (ISBLANK)
    side2 = Side(border_style="thin", color='2f3542')
    kwds = dict(fill_type="lightUp", start_color='BFBFBF')
    kwds.update(dict(product(sides, [side2])))
    CellFormats["isblank"] = EXCEL.Formulaformat(kwds)

    # Fill pattern for normal columns
    side4 = Side(border_style="thin", color='2f3542')
    kwds = dict(fill_type="solid", start_color='D0CECE', 
                size=9, wrap_text=True, color="404040", 
                vertical="center", horizontal="center", 
                bold=True)
    kwds.update(dict(product(sides,[side4])))
    CellFormats["columns"] = EXCEL.CellStyle(kwds)

    # Fill pattern of added columns
    ColumnFormats = dict()
    for key in params.hex.keys():
        kwds.update({k: params.hex[key][k] for k in ["start_color", "color"]})
        ColumnFormats[key] = EXCEL.CellStyle(kwds)
    
    Params = namedtuple("Params", ("cell","cf"))
    return Params(CellFormats, ColumnFormats)

def AuditSheet(sh, params):
    
    '''Create "Audit" sheet'''
    # Add new columns to "Audit" sheet
    new_columns = [c[1] for c in params.cols["Audit"]]
    max_columns = sh.max_column + 1
    all_columns = [c.value for c in sh[1]] + new_columns
    for c,col in enumerate(new_columns, max_columns):
        sh.cell(1,c).value = col

    # Adding ref_id
    hashcol = [c[1] for c in params.cols['hash']]
    hashidx = [n for n,c in enumerate(sh[1],1) if c.value in hashcol]
    ref_col = np.argmax(np.isin(all_columns, "ref_id")) + 1
    emp_col = np.argmax(np.isin(all_columns, "emp_id")) + 1
    for r in range(2, sh.max_row + 1):
        raw = "".join([str(v) for c in hashidx 
                       if (v:=sh.cell(r,c).value) is not None])
        sh.cell(r, ref_col).value = hmac_hash(raw)
    
    # Apply format to "Audit" sheet                
    ApplyFormats(sh, 
                 CellFormats=params.cell, 
                 ColumnFormats=params.cf, 
                 DataValidations=params.dv)

    # Turn grid lines off
    sh.sheet_view.showGridLines = False

    # Apply autofilter
    EXCEL.AutoFilter(sh, header=1)

    # Protect columns
    new_columns.remove("ref_id")
    EXCEL.ProtectSheet(sh, labels=new_columns, 
                       password="admin", 
                       protect=False)

def ResultSheet(sh1, sh2, params):
    
    '''Create "Result" sheet'''
    # Add new columns to "Result" sheet
    for c,col in params.cols["Result"]:
        sh2.cell(1,c).value = col
    
    # Find position of ref_id and emp_id
    all_columns = [c.value for c in sh1[1]] + \
    [c[1] for c in params.cols["Audit"]]
    ref_col = np.argmax(np.isin(all_columns, "ref_id")) + 1
    emp_col = np.argmax(np.isin(all_columns, "emp_id")) + 1
    
    # Add "ref_id" and "emp_id"
    for r in range(2, sh1.max_row + 1):
        sh2.cell(r, 1).value = sh1.cell(r, ref_col).value
        sh2.cell(r, 2).value = sh1.cell(r, emp_col).value

    # Apply format to "Audit" sheet                
    ApplyFormats(sh2, 
                 CellFormats=params.cell, 
                 ColumnFormats=params.cf, 
                 DataValidations=params.dv)

    # Turn grid lines off
    sh2.sheet_view.showGridLines = False

    # Apply autofilter
    EXCEL.AutoFilter(sh2, header=1)

def find_lastrow(sh, delete=False):
    
    '''Find the last row that contains data'''
    row = sh.max_row
    empty = True
    while empty:
        column = 1
        while column<=sh.max_column:
            if sh.cell(row,column).value is not None: 
                empty = False; break
            else: column += 1
        else: row -= 1
        if delete: sh.delete_rows(row+1)    
    return row

def find_lastcolumn(sh, delete=False):
    
    '''Find the last column that contains data'''
    column = sh.max_column
    empty = True
    while empty:
        row = 1
        while row<=sh.max_row:
            if sh.cell(row,column).value is not None: 
                empty = False; break
            else: row += 1
        else: column -= 1
        if delete: sh.delete_cols(column+1)    
    return column

class AuditReport_base:
    
    '''
    Create Internal Fraud report.
    
    Parameters
    ----------
    source : str
        Valid configuration file path e.g. "C:\\Folder\\exmaple.xlsx"
    
    sheetname : str
        Name of worksheet that contains configurations.
        
    '''
    def __init__(self, source, sheetname):
        self.source = source
        wb = load_workbook(source)
        self.sheetnames = ["Audit", "Result"]
        p0 = GetParameters(wb[sheetname], self.sheetnames)
        p1 = GeneralFormats(p0)
        self.params = sum_namedtuple("Params").merge(p0,p1)
        wb.close()
        self.Workbook = openpyxl.workbook.workbook.Workbook
    
    def create(self, source, saveas=None):
        
        '''
        Create "Audit", and "Result" sheets.
        
        Parameters
        ----------
        source : str or openpyxl workbook
            A valid source file path or openpyxl workbook object.
        
        saveas : str
            A valid repository file path e.g. "C:\\Folder\\exmaple.xlsx".
            If None or invalid file path, it defualts to current working 
            directory with defualt file name as a destination i.e. 
            "AUDIT_%Y%m%d%H%M%S.xlsx".
            
        '''
        # Validate source
        if isinstance(source, str): 
            if os.path.exists(source):
                wb = load_workbook(source)
            else: raise ValueError(f"Invalid file path: {source}.")
        elif not isinstance(source, self.Workbook):
            raise ValueError(f"source must be either str or "
                             f"openpyxl workbook. Got {type(source)}.")
        
        # Validate save as
        dirname = os.path.dirname(str(saveas))
        if (((os.path.exists(dirname)==False)) 
            and (dirname!="") or (saveas is None)): 
            saveas = f"AUDIT_{datetime.now():%Y%m%d%H%M%S}.xlsx"
    
        # Create sheets i.e. "Audit", and "Result"
        wb  = EXCEL.CopySheet(wb, "Data", self.sheetnames[0])
        sh1 = wb[self.sheetnames[0]]
        sh2 = wb.create_sheet(self.sheetnames[1], 1) 
        row = find_lastrow(sh1, True)
        col = find_lastcolumn(sh1, True)
        AuditSheet(sh1, self.params)
        ResultSheet(sh1, sh2, self.params)
        wb.save(saveas)
        wb.close()
        del wb

def findword(s, regex):
    return p[0] if len(p:=re.findall(regex,s))>0 else None

def SaveExcel(attachments, workbook, saveas, fields, sheetname="Data"):
    
    '''
    Save attchments (*.xls).
    
    Parameters
    ----------
    attachments : dict
         Key represents file name, and value contains is the file e.g.
         {'example.xlsx': <COMObject <unknown>>}.
    
    workbook : str
        Name of workbook to be saved.
    
    saveas : str
        A valid repository file path e.g. "C:\\Folder\\exmaple.xlsx".
        If invalid, it creates a directory recursively. That means if 
        any intermediate-level directory is missing, os.makedirs() 
        method will create them all.
        
    fields : list of str
        A list of madatory column labels.
        
    sheetname : str, default="Data"
        String that is used for sheet name.
    
    Returns
    -------
    content : dictionary
    
    '''
    if os.path.exists(saveas)==False: os.makedirs(saveas)
    path = os.path.join(saveas, workbook)
    keys = ["sheets", "n_sheet", "hasdata", "n_rows", 
            "n_cols", "range", "n_points", "p_point", 
            "p_match", "username", "computer"]
    content = dict([(key, None) for key in keys])
    for name,obj in attachments.items():
        if os.path.splitext(name)[1].find(".xls")>-1:
            obj.SaveAsFile(path)
            wb = load_workbook(path)
            content = {"sheets"  : ",".join(wb.sheetnames),
                       "n_sheets": len(sheets:=wb.sheetnames),
                       "hasdata" : (hasdata:=sheetname in sheets),
                       "n_rows"  : (nr:=wb[sheetname].max_row if hasdata else 0),
                       "n_cols"  : (nc:=wb[sheetname].max_column if hasdata else 0), 
                       "range"   : (rng:=f"A2:{get_column_letter(max(nc,1))}{max(nr,2)}"),
                       "n_points": (n_points:=len([n.value 
                                                   for n in np.r_[wb[sheetname][rng]] 
                                                   if n.value is not None]) 
                                    if hasdata else 0), 
                       "p_point" : round(100*n_points/max((nr-1)*nc,1),2), 
                       "p_match" : round(100*len([c.value for c in wb[sheetname]["1"] 
                                                  if c.value in fields])/len(fields) 
                                         if hasdata else 0,2),
                       "username": os.environ["USERNAME"],
                       "computer": os.environ["COMPUTERNAME"],
                       "complete": "FALSE",
                       "SendTime": None, 
                       "SendUser": None,
                       "SendComputer": None}
            wb.close()
            del wb
    return content

def ExtractMailContents(ReportGenerator, mails, source):
    
    '''
    Extract contents from mails as well as save attachments (if any)
    ReportGenerator : AuditReport_base class
    mails : output from OUTLOOK.ReadMail
    source : valid file path for repository
    '''
    # Initialize widgets, and paramters
    t = widgets.HTMLMath(value='')
    display(widgets.HBox([t]))
    n_valid = 0
    files = dict()
    
    # Mail subject content (lower case)
    subj = "Internal fraud Data".lower()
    
    # Mandatory fields
    fields = [c for _,c in ReportGenerator.params.cols["mandate"]]
    logger.info(f"์Number of emails : {len(mails):,.0f}.")

    for n in tqdm_notebook(mails.keys()):
        mail = namedtuple("Mail", mails[n].keys())(*mails[n].values())
        t.value = mail.Subject

        if (att:=mail.Attachments) is not None:
            if re.findall(subj, mail.Subject.lower()).count(subj)>0:
                
                text = "\n".join([mail.Subject, mail.Body])
                content = {"period"       : (prd:=findword(text, "[0-9]{8}")),
                           "pattern"      : (pat:=findword(text, "[A-Z][0-9]{3}")),
                           "sender"       : mail.SenderName, 
                           "to"           : mail.To, 
                           "cc"           : mail.CC, 
                           "Subject"      : mail.Subject, 
                           "Body"         : mail.Body.replace("\r","").replace("\n",""), 
                           "ReceivedTime" : mail.ReceivedTime, 
                           "CreationTime" : mail.CreationTime, 
                           "SentOn"       : mail.SentOn,
                           "UpdateTime"   : f"{datetime.now():%d-%m-%Y %H:%M:%S}",
                           "workbook"     : (workbook:=f"{prd}_{pat}.xlsx"), 
                           "saveas"       : (saveas:=os.path.join(source, str(pat))), 
                           "NoAttachment" : len(att.keys()), 
                           "Attachments"  : ",".join(list(att.keys()))}

                content.update(SaveExcel(att, workbook, saveas, fields))
                files[n_valid] = content
                new_source = os.path.join(saveas, workbook)
                ReportGenerator.create(new_source, new_source)
                n_valid += 1
                logger.debug(f"<{workbook}> has been created successfully.")
                
            else: logger.warning(f"<{mail.Subject}> received at "
                                 f"<{mail.ReceivedTime}> is not "
                                 f"an internal fraud data.")
                
        else: logger.warning(f"<{mail.Subject}> received at "
                             f"<{mail.ReceivedTime}> does not "
                             f"have attachment(s).")
                
    t.value = "# of emails w/ attachments : {:,.0f}".format(n_valid)
    return files

def ApplySummaryFormats(path):
    
    '''Apply formats to sheet, namely "Summary".'''
    # Initialize parameters
    wb = load_workbook(path)
    sh = wb["Summary"]
    sides = ["right", "left", "top", "bottom"]
    columns = [c.value for c in sh["1"]]
    offset = (sh.max_row-1, sh.max_column-1)

    # General format
    side1 = Side(border_style="thin", color='2f3542')
    side1 = dict(product(sides, [side1]))
    EXCEL.CellStyle(side1).apply(sh, offset=offset)

    # Format for cells with data validation (ISBLANK)
    side2 = Side(border_style="thin", color='2f3542')
    kwds = dict(fill_type="lightUp", start_color='BFBFBF')
    kwds.update(dict(product(sides, [side2])))
    EXCEL.Formulaformat(kwds).apply(sh, offset=offset)

    # Fill pattern for normal columns
    side4 = Side(border_style="thin", color='2f3542')
    kwds = dict(fill_type="solid", start_color='D0CECE', 
                size=9, wrap_text=True, color="404040", 
                vertical="center", horizontal="center", 
                bold=True)
    kwds.update(dict(product(sides,[side4])))
    EXCEL.CellStyle(kwds).apply(sh, offset=(0, sh.max_column-1))

    # Format for cells with data validation (ISBLANK)
    side2 = Side(border_style="thin", color='2f3542')
    kwds = dict(fill_type="lightUp", start_color='FF0000')
    kwds.update(dict(product(sides, [side2])))
    ErrorFormat = EXCEL.Formulaformat(kwds)
    Errorparams = {'hasdata': '{}2=FALSE'.format, 
                   'n_rows' : '{}2=0'.format, 
                   'n_cols' : '{}2<12'.format,
                   'p_match': '{}2<100'.format}
    
    # Apply formats to cells
    for col,formula in Errorparams.items():
        c = np.argmax(np.isin(columns,col))+1 
        ErrorFormat.apply(sh, (2, c), offset=(sh.max_row-2, 0), 
                          formula=formula(get_column_letter(c)))
        
    # Apply validation
    c = np.argmax(np.isin(columns,"complete"))+1 
    kwargs = {"type" : "list", "formula1" : "TRUE, FALSE",}
    EXCEL.Validation(**kwargs).apply(sh, (2,c), offset=(sh.max_row-2, 0))

     # Turn grid lines off
    sh.sheet_view.showGridLines = False

    # Apply autofilter
    EXCEL.AutoFilter(sh, header=1)
    
    wb.save(path)
    wb.close()

class AuditReport:
    
    '''
    Create internal fraud report (*.xlsx)
    
    Parameters
    ----------
    source1 : str
        A valid excel file (*.xlsx) path for predefined settings.
        
    sheetname : str, default="Settings"
        Name of worksheet in "source1" that contains all settings.
    
    stop : str, default=None
        A stop date (format = "%d/%m/%Y %H:%M:%S"). If None, it uses 
        the current date i.e. datetime.now().
        
    start : str, default=None
        A start date (format = "%d/%m/%Y %H:%M:%S"). If None, it uses 
        the latest "ReceivedTime" from "source2". If "source2" is None, 
        it uses "01/01/2022 00:00:00".
        
    source2 : str, default=None
        A valid excel file (*.xlsx) path for previously received emails.
        If None, it searches under current working directory otherwise
        it assumes no recieved emails and create a new file under same
        directory.
    
    source3 : str, default=None
        A valid folder path for storing outputs. if None, a default 
        folder namely "data", is created under current working directory.
    
    Attributes
    ----------
    mails : dict
        For each key, it contains dictionary, whose keys (properties) are 
        as follows:
        - "SenderName"   : Name of the sender
        - "To"           : List of recipients
        - "CC"           : List of carbon copy (CC) names
        - "Subject"      : Mail subject
        - "Body"         : The clear-text body of the Outlook item
        - "HTMLBody"     : HTML body of the specified item
        - "ReceivedTime" : Date & time at which the item was received
        - "CreationTime" : Creation time
        - "SentOn"       : Date & time at which the item was sent
        - "Attachments"  : Attached files.
        - "sheets"       : Worksheet names
        - "n_sheets"     : Number of worksheets
        - "hasdata"      : If TRUE, it contains "Data" worksheet
        - "n_rows"       : Number of rows that contain data
        - "n_cols"       : Number of columns that contain data
        - "range"        : Data range (address)
        - "n_points"     : Number of data points
        - "p_point"      : % data point
        - "p_match"      : % match with mandatory columns
        - "username"     : User name
        - "computer"     : Computer name
        - "complete"     : FALSE (Whether the data is complete or not)
        - "SendTime"     : None (Send time)
        - "SendUser"     : None (Send user)
        - "SendComputer" : None (Send computer)
        
    '''
    datetime_format = "%d/%m/%Y %H:%M:%S"
    min_date = datetime.strptime("01/01/2022 00:00:00", datetime_format)
    str_date = "{:%d/%m/%Y %H:%M:%S}".format
    sum_file  = "InternalFraud.xlsx"
    
    def __init__(self,  source1:str, 
                 sheetname:str="Settings", 
                 stop:str=None, 
                 start:str=None, 
                 source2:str=None, 
                 source3:str=None):
        
        # Default path under current working directory.
        defpath = os.path.join(os.getcwd(),"data")
        self.source1 = source1
        self.source2 = (os.path.join(defpath, self.sum_file)  
                        if source2 is None else source2)
        self.source3 = defpath if source3 is None else source3
        self.sheetname = sheetname
  
        self.kwargs = {"start" : (self.__lastestdate__(self.source2) 
                                  if start is None else start), 
                       "stop"  : (self.str_date(datetime.now()) 
                                  if stop is None else stop),
                       "sort"  : False}
        self.generator = AuditReport_base(self.source1, 
                                          self.sheetname)
        
    def __lastestdate__(self, path:str):
        
        '''Get the latest ReceivedTime'''
        self.prevmails = None
        if os.path.exists(path):
            self.prevmails = pd.read_excel(path, sheet_name="Summary")
            max_dt = pd.to_datetime(self.prevmails["ReceivedTime"],
                                    format=self.datetime_format).max() 
            max_dt = np.fmax(max_dt+timedelta(seconds=1), self.min_date)
            return self.str_date(max_dt)
        else: return self.str_date(self.min_date)

    def extract(self, source):
        
        '''
        Read and extract Microsoft Outlook mails from designated folder.
        
        Parameters
        ----------
        source : str
            A valid Microsoft Outlook folder path e.g. "xxx@gmail\Inbox".

        '''
        logger.info("Initialize [Creating Report]")
        mails = OUTLOOK.ReadMail(source, **self.kwargs)
        args  = (self.generator, mails, self.source3)
        self.mails = ExtractMailContents(*args)
        self.__summary__()
        logger.info("Terminate [Creating Report]")
        logging.shutdown()
        return self
    
    def __summary__(self):
        
        '''
        Create "Summary" worksheet.
        '''
        if len(self.mails)>0:
            df = pd.DataFrame(self.mails).T
            if self.prevmails is not None:
                df = pd.concat((self.prevmails, df), ignore_index=True)
            df.to_excel(self.source2, sheet_name="Summary", index=False)
            ApplySummaryFormats(self.source2)
            logger.debug(f"<{self.source2}> has been saved successfully.")

class SendReport:
    
    '''
    Copy excel files (report) to destination.
    
    Parameters
    ----------
    source : str
        A valid excel file (*.xlsx) path for predefined settings.
        
    sheet_name1 : str, default="Rename"
        Name of worksheet that contains renamed folders.
        
    sheet_name2 : str, default="Recipients"
        Name of worksheet in "source" that contains list of recipients.
        
    '''
    sheet_name = "Summary"
    usecols = ["SendTime","SendUser","SendComputer"]
    
    def __init__(self, source, 
                 sheet_name1="Rename", 
                 sheet_name2="Recipients", 
                 overwrite=True):
        
        self.overwrite = overwrite
        if not os.path.exists(source):
            raise ValueError(f'Invalid file path : "{source}".')
        
        # Create "rename" dictionary
        rename = pd.read_excel(source, sheet_name=sheet_name1)
        self.rename = dict([n for n in rename.values.tolist()])
        
        # Create list of recipients
        recp = pd.read_excel(source, sheet_name=sheet_name2)
        recp = recp.groupby("Send").agg({"Email":(lambda x: ";".join(x))})
        self.recipients = recp.to_dict()["Email"]
        self.kwargs = {"recipients" : self.recipients["to"].split(";"),
                       "cc"         : self.recipients["cc"].split(";"), 
                       "subject"    : "Internal Fraud Monitoring Team",
                       "display"    : True,
                       "kwargs"     : {"Importance" : 2, 
                                       "Sensitivity": 1}}

    def copy(self, source, destination=None, send=False, display=True):
        
        '''
        Copy files from "source" to "destination".

        Parameters
        ----------
        source : str
            A valid excel file (*.xlsx) path of summary report generated
            by AuditReport (class).
            
        destination : str, default=None
            A valid destination path (folder) where files are copied to.
            If None, a default folder is created i.e. NEW_%Y%m%d%H%M%S
            under current working directory.
            
        send : bool, default=False
            If True, it sends the notification email according to email
            list.
            
        display : bool, default=True
            If True, it displays the mail, otherwise sends. This is
            relevant when "send" is True.
            
        '''
        if destination is None:
            new_folder = f"NEW_{datetime.now():%Y%m%d%H%M%S}"
            self.destination = os.path.join(os.getcwd(), new_folder)
        else: self.destination = destination
            
        logger.info("Initialize [Sending Files]")
        logger.info(f"Main folder : <{self.destination}>.")
            
        if os.path.exists(self.destination)==False: 
            os.makedirs(self.destination)
            
        # Every file that has not been sent out.
        locales = pd.read_excel(source, sheet_name=self.sheet_name)
        unsents = locales.loc[locales["SendTime"].isna() & 
                              (locales["complete"]==True),
                              ["saveas","workbook","pattern"]]
        unsents = unsents.reset_index().values
        logger.info("Number of unsent files : {:,.0f}.".format(len(unsents)))
        
        # Copy file from source to destination
        for n, saveas, workbook, pattern in tqdm_notebook(unsents):
            
            locale0 = os.path.join(saveas, workbook)
            locale1 = os.path.join(destination, self.rename[pattern], workbook)
            success = True
            
            if os.path.exists(locale0):
     
                folder, file = os.path.split(locale1)
                if not os.path.exists(folder):
                    os.makedirs(folder)
                    folder = folder.replace(destination,"")
                    logger.debug(f"<..{folder}> has been created successfully.")
                    
                # Copy file to destination
                if os.path.exists(locale1) and self.overwrite:
                    dest = shutil.copyfile(locale0, locale1)
                    logger.debug(f"<{file}> has been overwritten successfully.")
                elif not os.path.exists(locale1):    
                    dest = shutil.copyfile(locale0, locale1)
                    logger.debug(f"<{file}> has been copied successfully.")
                else: 
                    logger.warning(f"Copy failed. <{file}> already exists.")
                    success = False
 
            else: logger.debug(f"<{locale0}> cannot be found.")
            
            # Update summary files
            if success:
                locales.loc[n, self.usecols] = [f"{datetime.now():%d-%m-%Y %H:%M:%S}",
                                                os.environ["USERNAME"],
                                                os.environ["COMPUTERNAME"]]
            
            # Sending out mail to notify relevant personnel
            if success & send:
                period = str(locales.loc[n,'period'])
                period = "{:%d-%m-%Y}".format(datetime.strptime(period,"%Y%m%d"))
                html = CreateHTML(self.rename[pattern], workbook, period)
                OUTLOOK.SendMail(html, **{**self.kwargs,**{"display":diplay}})
                logger.debug(f"์Notification email of <{workbook}> "
                             f"has been sent out successfully.")
        
        # Save updated summary (*.xlsx) and apply format
        locales.to_excel(source, sheet_name="Summary", index=False)
        ApplySummaryFormats(source)
        logger.debug(f"<{source}> has been saved successfully.")
        logger.info("Terminate [Sending Files]")
        logging.shutdown()
        
        return self

def CreateHTML(folder, file, period):
    '''Create html of outlook email'''
    link  = ("https://kasikornbankgroup-my.sharepoint.com/personal/piti_p_kasikornbank_com/"
             "_layouts/15/onedrive.aspx?FolderCTID=0x012000FDB26412F17E3A4FB421F4C3C7F609F8&id="
             "%2Fpersonal%2Fpiti%5Fp%5Fkasikornbank%5Fcom%2FDocuments%2Fwork%2Finternal%5Ffraud"
             "%5Fmonitoring%2FFR%5FMonitoringCenter%2F")
    color, font, size = "Black", "Tahoma", "13px"
    font0 = f'<p style=""font-family:{font}; color:{color}; font-size: {size};"">'
    font1 = f'<p style=""font-family:{font}; color:{color}; font-size: {size}; text-indent: 30px"">'
    link += urllib.parse.quote(folder) + "&view=0"
    html  = [f"{font0}<b>Internal Fraud Monitoring Team</b> </p>",
             f"{font1} &emsp; นำส่งไฟล์ Pattern รอบ {period} ตามรายละเอียดด้านล่าง</p>", 
             f'{font1} &emsp; Pattern : <a href={link}>{folder}</a> </p>', 
             f"{font1} &emsp; File name : {file}</p>", 
             f"{font0} <b>Best Regards</b>,</p>",
             f"{font0} Fraud Analytics"]
    return "<!DOCTYPE html><HTML><body>{}</body></HTML>".format("".join(html))