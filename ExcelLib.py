'''
Available methods are the followings:
[1] AutoFilter
[2] ProtectSheet
[3] Validation (Class)
[4] CellStyle (Class)
[5] CopySheet
[6] get_filepaths

Authors: Danusorn Sitdhirasdr <danusorn.si@gmail.com>
versionadded:: 11-07-2023

'''
import pandas as pd, numpy as np, os
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import (PatternFill, Border, Side, 
                             Alignment, Font, Protection)
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting import Rule
from openpyxl.utils import get_column_letter
from itertools import product

__all__ = ["AutoFilter",
           "ProtectSheet", 
           "Validation",
           "CellStyle", 
           "CopySheet", 
           "get_filepaths"]

class Validation:
    
    '''
    Add validation to range.
    
    Parameters
    ----------
    type : str, default=None
        Specify type of validation. The supported types are as follows:
        "whole"      : to restrict cell to accept only whole numbers.
        "decimal"    : to restrict cell to accept only decimal numbers.
        "list"       : to pick data from the drop-down list.
        "date"       : to restrict cell to accept only date.
        "time"       : to restrict cell to accept only time.
        "textlength" : to restrict the length of the text.
        "custom"     : for custom formula.

    operator : str, default=None
        Specify the operator. This is relevant when type is not "list"
        or "customr". The supported operators are "lessThanOrEqual", 
        "equal", "greaterThan", "notEqual", "lessThan", "between", 
        "greaterThanOrEqual", and "notBetween".
    
    formula1 : str, int, float, default=None
        Valid input for cell e.g. "a,b,c" or 100 or "=ISNUMBER(A1)" . 
        If operator is either "between" or "notBetween", formula1 
        becomes a lower bound. 
        
        Input for date and time "%d/%m/%Y"
        
        
    formula2 : str, int, float, default=None
        An upper bound when operator is either "between" or 
        "notBetween". 
        
    allow_blank : bool, default=True
        If True, it allows cell to be blank.
        
    kwargs: dict, default=None
        Other attributes e.g. 'error', 'errortitle', 'prompt', 
        and 'prompttitle'.
        
    References
    ----------
    [1] https://openpyxl.readthedocs.io/en/stable/styles.html
    [2] https://openpyxl.readthedocs.io/en/latest/validation.html
    [3] https://support.microsoft.com/en-us/office/apply-data-
        validation-to-cells-29fecbcc-d1b9-42c1-9d76-eff3ce5f7249
        
    '''
    def __init__(self, type=None, operator=None, formula1=None, 
                 formula2=None, allow_blank=True, kwargs=None):
        
        # Initialize parameters
        self.input = {"date" : ("31/12/1899", "%d/%m/%Y", "days", 1), 
                      "time" : ("00:00:00", "%H:%M:%S", "seconds", 1/(24*3600))}
        
        # Convert inputs for date & time
        if type == "date":
            formula1 = int(FindDiff(*((formula1,) + self.input[type]))) + 1
            formula2 = int(FindDiff(*((formula2,) + self.input[type]))) + 1
        if type == "time":
            formula1 = round(FindDiff(*((formula1,) + self.input[type])), 15)
            formula2 = round(FindDiff(*((formula2,) + self.input[type])), 15)
        elif type=="list": 
            formula1 = formula1[:255].lstrip().rstrip()
            formula1 =f'"{formula1}"'
        elif type=="custom":
            formula1 =f'"{formula1}"'
        else: pass
        
        # Create data validation
        self.dv = DataValidation(type=type, 
                                 operator=operator,
                                 formula1=formula1,
                                 formula2=formula2,
                                 allow_blank=allow_blank)
            
        # Optionally set a custom message
        if isinstance(kwargs, dict):
            for key in kwargs.keys():
                setattr(self.dv, key, kwargs[key])

    def apply(self, sh, coord=(1,1), offset=(0,0)):
        
        '''
        Transform designated cell(s) according to validation format.
        
        Parameters
        ----------
        sh : openpyxl Worksheet
            Excel worksheet.

        coord : (int,int), default=(1,1)
            Starting cell coordinate e.g. (1,1) is "A1".

        offset : (int,int), default=(0,0)
            The number of rows, and columns to be added.
            
        '''
        # Add the data-validation object to designated cell
        cell0 = sh.cell(*coord).coordinate
        cell1 = sh.cell(coord[0] + max(offset[0], 0), 
                        coord[1] + max(offset[1], 0)).coordinate
        
        # Remove previously added range(s)
        while len(self.dv.ranges.ranges):
            rng = self.dv.ranges.ranges[0].__str__()
            self.dv.ranges.remove(rng)
        
        # Add new range
        self.dv.add(f"{cell0}:{cell1}")
        sh.add_data_validation(self.dv)

def FindDiff(dt0, dt1, dtformat, attr, factor=1):
    '''Find absolute difference between dt0 and dt1'''
    if dt0 is not None:
        diff = (datetime.strptime(dt0, dtformat)-
                datetime.strptime(dt1, dtformat))
        try: return abs(getattr(diff, attr, None)) * factor
        except: return None
    else: return None

def UpdateDict(dict1:dict, dict2:dict):
    '''Update dictionary'''
    if isinstance(dict1, dict) & isinstance(dict2, dict):
        for key in set(dict1.keys()).intersection(dict2.keys()):
            dict1[key] = dict2[key]
    return dict1

def CreateCoords(coord=(1,1), offset=(0,0)):
    '''Create coordinates'''
    stop = tuple(np.r_[coord] + offset + (1,1))
    return list(product(range(coord[0], stop[0]),
                        range(coord[1], stop[1])))

class CellStyle:
    
    '''
    Change cell format.
    
    Parameters
    ----------
    kwargs: dict, default=None
        Other parameters of the following openpyxl.styles functions:
        - PatternFill e.g. "fill_type", "start_color"
        - Alignment e.g. "horizontal", "vertical"
        - Font e.g. "name", "underline"
        - Border e.g. "left", "right", whose value must be 
          "openpyxl.styles.Side(border_style, color)"
    
    References
    ----------
    [1] https://openpyxl.readthedocs.io/en/stable/styles.html
    
    '''
    def __init__(self, kwargs=None):
        
        # Default properties
        # underline = {'single', 'double', 
        #              'doubleAccounting', 
        #              'singleAccounting'}
        self.fontstyle = dict(name='Tahoma', 
                              size=10, 
                              bold=False, 
                              italic=False, 
                              vertAlign=None, 
                              underline='none', 
                              strike=False, 
                              color='404040')

        self.fillstyle = dict(fill_type="none", 
                              start_color='000000', 
                              end_color='000000', 
                              patternType='none')
        
        self.alignment = dict(horizontal='general',
                              vertical='bottom',
                              text_rotation=0,                 
                              wrap_text=False,
                              shrink_to_fit=False,                    
                              indent=0)
        
        # vertical and horizontal are excluded
        # same parameters as "alignment"
        sidestyle = Side(border_style=None, color='000000')
        self.borderstyle = dict(left=sidestyle,
                                right=sidestyle,
                                top=sidestyle,
                                bottom=sidestyle,
                                diagonal=sidestyle,
                                diagonal_direction=0,
                                outline=sidestyle)
        
        kwargs = dict() if kwargs is None else kwargs
        self.fontstyle = Font(**UpdateDict(self.fontstyle, kwargs))
        self.fillstyle = PatternFill(**UpdateDict(self.fillstyle, kwargs))
        self.alignment = Alignment(**UpdateDict(self.alignment, kwargs))
        self.borderstyle = Border(**UpdateDict(self.borderstyle, kwargs))
    
    def apply(self, sh, coord=(1,1), offset=(0,0)):
        
        '''
        Apply predefined format to designated cell(s).
        
        Parameters
        ----------
        sh : openpyxl Worksheet
            Excel worksheet.

        coord : (int,int), default=(1,1)
            Starting cell coordinate e.g. (1,1) is "A1".

        offset : (int,int), default=(0,0)
            The number of rows, and columns to be added.
            
        '''
        for cell in CreateCoords(coord, offset):
            addr = sh.cell(*cell).coordinate
            sh[addr].font = self.fontstyle
            sh[addr].fill = self.fillstyle
            sh[addr].alignment = self.alignment
            sh[addr].border = self.borderstyle

def AutoFilter(sh, header=1, start=1, end=None):
    
    '''
    Add auto filter.
    
    Parameters
    ----------
    sh : openpyxl Worksheet
        Excel worksheet.
    
    header : int, default=1
        Row to use for columns.
        
    start : int, default=1
        Starting column index.
    
    end : int, default=None
        Ending column index. If None, it uses maximum number of
        used column range (max_column)
        
    '''
    start = int(np.fmax(1, start))
    end = end if isinstance(end,int) else sh.max_column
    end = int(np.fmax(np.fmin(end, sh.max_column), start))
    filters = sh.auto_filter
    filters.ref = "{}{}:{}{}".format(get_column_letter(start), header,
                                     get_column_letter(end), sh.max_row)

def ProtectSheet(sh, labels=None, header=1, password="admin", protect=True):
    
    '''
    Protect columns in worksheet.
    
    Parameters
    ----------
    sh : openpyxl Worksheet
        Excel worksheet.
        
    labels : list of str, default=None
        List of column labels to be protected. If None, it protects
        all columns.
    
    header : int, default=1
        Row to use for the column labels.

    password : str
        Password to protect worksheet.
  
    protect : bool, default=True
        If False, it protects all columns not in list. If True, it 
        protects all columns from the list.
        
    '''
    # Protected columns
    columns = [c.value for c in sh[str(header)]]
    labels = columns if labels is None else labels
    if protect==False: 
        labels = list(set(columns).difference(labels))

    # Protect worksheet
    sh.protection.sheet = False
    
    # Protect / Unprotect columns
    for c,col in enumerate(columns,1):
        locked = True if col in labels else False
        sh.cell(1,c).protection = Protection(locked=True)
        for r in range(header + 1, sh.max_row+1):
            sh.cell(r,c).protection = Protection(locked=locked)
    
    # Protect sheet
    sh.protection = SheetProtection(sh, 
                                    formatRows=False, 
                                    formatColumns=False, 
                                    selectLockedCells=False, 
                                    selectUnlockedCells=False, 
                                    autoFilter=False,
                                    sort=False,
                                    password=password)
    
    # Save and close workbook
    sh.sheet_view.showGridLines = False

def CopySheet(wb, sheetname, new_sheetname=None, new_wb=None):
    
    '''
    Copy worksheet to new worksheet.
    
    Parameters
    ----------
    wb : openpyxl Workbook
        Workbook.
        
    sheetname : str
        Name of worksheet to be copied.
    
    new_sheetname : str, default=None
        Name for new worksheet. If None or same name is found, it 
        defaults to "NEW_%Y%m%d%H%M%S".
        
    new_wb : openpyxl Workbook
        New workbook. If None, it uses current workbook ("wb").
        
    References
    ----------
    [1] https://openpyxl.readthedocs.io/en/stable/tutorial.html
    
    Returns
    -------
    new_wb : openpyxl Workbook
        New workbook with new worksheet.
    
    '''
    # Initiate parameters
    def_name = f"NEW_{datetime.now():%Y%m%d%H%M%S}"
    sh = wb[sheetname]
    new_wb = wb if new_wb is None else new_wb
    if ((new_sheetname is None) or 
        (new_sheetname in new_wb.sheetnames)): 
        new_sheetname = def_name
    new_sh = new_wb.create_sheet(new_sheetname, 0)

    # copying the cell values from source 
    # excel file to destination excel file
    for i in range (1, sh.max_row + 1):
        for j in range (1, sh.max_column + 1):
            # writing the read value to destination excel file
            new_sh.cell(i,j).value = sh.cell(i, j).value

    return new_wb

def GetCellRange(sh, coord=(1,1), offset=(0,0)):
    '''Return address of cell range'''
    start = sh.cell(*coord).coordinate
    stop = tuple(np.r_[coord] + offset)
    stop = sh.cell(*stop).coordinate
    return start, stop

class Formulaformat(CellStyle):
    
    '''
    Change cell format given formula.
    
    Parameters
    ----------
    kwargs: dict, default=None
        Other parameters of the following openpyxl.styles functions:
        - PatternFill e.g. "fill_type", "start_color"
        - Alignment e.g. "horizontal", "vertical"
        - Font e.g. "name", "underline"
        - Border e.g. "left", "right", whose value must be 
          "openpyxl.styles.Side(border_style, color)"
    
    References
    ----------
    [1] https://openpyxl.readthedocs.io/en/stable/formatting.html
    
    '''
    def __init__(self, kwargs=None):
        
        # alignment is excluded (not compatible)
        self.pattern = CellStyle(kwargs)
        self.params = {"font"   : self.pattern.fontstyle, 
                       "fill"   : self.pattern.fillstyle, 
                       "border" : self.pattern.borderstyle}

    def apply(self, sh, coord=(1,1), offset=(0,0), formula=None):
        
        '''
        Add conditional format (blank) to designated cell(s).
        
        Parameters
        ----------
        sh : openpyxl Worksheet
            Excel worksheet.

        coord : (int,int), default=(1,1)
            Starting cell coordinate e.g. (1,1) is "A1".

        offset : (int,int), default=(0,0)
            The number of rows, and columns to be added.
            
        formula : str, default=None
            Rule formula. If None, "ISBLANK(cell.address)" is used.
          
        '''
        addr = GetCellRange(sh, coord, offset)
        if formula is None: formula = f'ISBLANK({addr[0]})'
        rule = FormulaRule(formula=[formula], 
                           stopIfTrue=True, **self.params)
        sh.conditional_formatting.add(":".join(addr), rule)

def get_filepaths(path=None):
    
    '''
    Find all file directories, which are scanned from top-down. 
    
    Parameters
    ----------
    path : str, default=None
        Root directory. If None, it defaults to current working 
        directory i.e. os.getcwd().
    
    Returns
    -------
    foundfiles : dict
        {"file1.xlsx" : {"path" : "C:\\file1.xlsx", 
                         "size" : 50000},
         "file2.xlsx" : {"path" : "C:\\file2.xlsx", 
                         "size" : 11000}}
    
    '''
    if path is None: path = os.getcwd()
    foundfiles = dict()
    for root, dirs, files in os.walk(path, topdown=True):
        for name in files:
            filename = os.path.join(root,name)
            foundfiles[name] ={"path" : filename, 
                               "size" : os.path.getsize(filename)}
    return foundfiles